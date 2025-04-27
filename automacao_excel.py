import openpyxl


# Criar uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Vendas"

# Adicionar Cabecalhos

ws['A1'] = "Descricao"
ws['B1'] = "Quantidade"
ws['C1'] = "Valor Unitario"
ws['D1'] = "Valor Total"

# Adicionar alguns dados

ws.append(["Abaixador de Lingua - Pacote com 100und", 15, 12.00, 12.00 * 15])
ws.append(["Cateter EV 25g - caixa com 100und", 4, 27.00, 27,00 * 4])
ws.append(["Agulha 30x0,07 cx 100und", 35, 24.00, 34,00 * 35])

# Ajustando a formatação para o valor unitário e total (com vírgula)
for row in range(2, ws.max_row + 1):
    ws[f'C{row}'].number_format = '#,##0.00'  # Formato monetário com vírgula (Valor Unitário)
    ws[f'D{row}'].number_format = '#,##0.00'  # Formato monetário com vírgula (Valor Total)

# Limpar possíveis células extras
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row:
        # Garantir que apenas os valores das 4 colunas sejam preenchidos
        if cell.value is None:
            cell.value = ""  # Apagar células vazias extras


# Salvar Arquivo
wb.save("relatorio_vendasatt1.xlsx")
print("planilha criada com sucesso.")

